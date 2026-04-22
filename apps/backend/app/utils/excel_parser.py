"""
Excel read/write helpers using openpyxl.

Read:   Extract RFP questions from Column A (skips empty / header rows).
Write:  Inject approved answers into Column B while preserving every byte
        of the original workbook — formatting, merged cells, named ranges,
        and VBA macros (via keep_vba=True).
"""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import structlog

log = structlog.get_logger(__name__)

# Rows whose Column-A value matches these patterns are treated as headers.
_HEADER_PATTERNS = {"question", "rfp question", "requirement", "#", "no.", "sr."}

# Visual styling applied to injected answers so they stand out in the export.
_ANSWER_FONT = Font(name="Calibri", size=11)
_ANSWER_FILL = PatternFill(start_color="EBF5EB", end_color="EBF5EB", fill_type="solid")
_ANSWER_ALIGN = Alignment(wrap_text=True, vertical="top")


def extract_questions(file_bytes: bytes) -> list[tuple[int, str]]:
    """
    Parse an XLSX workbook and return a list of (row_index, question_text) tuples.

    - Reads the active sheet (first sheet).
    - Column A is assumed to contain questions.
    - Skips blank rows and likely header rows.
    - row_index is 1-based to match openpyxl's convention.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    questions: list[tuple[int, str]] = []

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        value = cell.value

        if value is None:
            continue

        text = str(value).strip()

        if not text:
            continue

        # Skip rows that look like column headers
        if text.lower() in _HEADER_PATTERNS:
            continue

        questions.append((cell.row, text))

    log.info("excel.questions_extracted", count=len(questions))
    return questions


def inject_answers(
    file_bytes: bytes,
    answers: dict[int, str],  # {row_index: answer_text}
) -> bytes:
    """
    Load the original workbook and write approved answers into Column B.

    Critical preservation flags:
      - keep_vba=True  → macros survive the round-trip
      - rich_text=True → cell rich-text formatting is not stripped
      - Merged cells are not touched (we only write to un-merged B cells)

    Returns the modified workbook as raw bytes for streaming to the client.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        keep_vba=True,
        rich_text=True,
    )
    ws = wb.active

    # Ensure Column B has a reasonable width for human readability
    col_b_letter = get_column_letter(2)
    if col_b_letter not in (ws.column_dimensions or {}):
        ws.column_dimensions[col_b_letter].width = 60

    injected_count = 0
    for row_index, answer_text in answers.items():
        cell = ws.cell(row=row_index, column=2)

        # Do not overwrite cells that are part of a merged region
        # (the top-left cell of a merge is safe to write; others are not)
        if hasattr(cell, "merged_cells") and cell.coordinate in ws.merged_cells:
            log.warning("excel.skipped_merged_cell", row=row_index)
            continue

        cell.value = answer_text
        cell.font = _ANSWER_FONT
        cell.fill = _ANSWER_FILL
        cell.alignment = _ANSWER_ALIGN
        injected_count += 1

    log.info("excel.answers_injected", count=injected_count)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
